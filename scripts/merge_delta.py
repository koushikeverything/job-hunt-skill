#!/usr/bin/env python3
"""Delta merge: load a prior master workbook, dedupe new findings against it,
append only genuinely new roles, mark expired rows, rebuild the workbook.

Usage:
  python merge_delta.py prior.xlsx new_roles.json out.xlsx [--expired ids.json]
  python merge_delta.py prior.xlsx --exclusions-only        # print org|title list for agent briefings

new_roles.json: same shape as build_master_xlsx.py input (may include leads/notes to append).
--expired: optional JSON list of {"organisation":..,"role_title":..,"reason":..} confirmed closed this run.
"""
import json, re, sys
from datetime import date
from openpyxl import load_workbook
sys.path.insert(0, __file__.rsplit("/", 1)[0])
from build_master_xlsx import build, ROLE_COLS, ROLE_KEYS  # noqa: E402


def norm(s):
    s = re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower())
    s = re.sub(r"\b(the|ltd|llp|pvt|inc|limited|private|foundation|india)\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def title_key(title):
    return " ".join([w for w in norm(title).split() if len(w) > 2][:4])


GENERIC = {"group", "services", "solutions", "global", "international", "trust", "society",
           "school", "college", "centre", "center", "capital", "finance", "financial", "bank",
           "via", "for", "and", "org", "ngo", "company", "consulting", "technologies", "tech"}


def org_words(org):
    # keep parenthetical text (often the acronym, e.g. "(CEE)") but flatten it
    org = re.sub(r"[()]", " ", str(org or ""))
    return [w for w in norm(org).split() if len(w) > 2]


def org_overlap(oa, ob):
    if oa[0] == ob[0]:
        return True
    ja, jb = " ".join(oa), " ".join(ob)
    if ja in jb or jb in ja:
        return True
    return bool((set(oa) - GENERIC) & (set(ob) - GENERIC))


def same_role(a, b):
    """Loose match: same title key and organisations share their leading word or one contains the other."""
    if title_key(a.get("role_title")) != title_key(b.get("role_title")):
        return False
    oa, ob = org_words(a.get("organisation")), org_words(b.get("organisation"))
    if not oa or not ob:
        return False
    return org_overlap(oa, ob)


def key(org, title):  # kept for --exclusions-only printing compatibility
    return " ".join(org_words(org)) + "|" + title_key(title)


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr = [str(h) for h in rows[0]]
    return [dict(zip(hdr, r)) for r in rows[1:] if any(r)]


def load_prior(path):
    wb = load_workbook(path)
    data = {"roles": [], "leads": [], "top_picks": [], "cv_keywords": [], "boards": [], "coverage": [], "notes": [], "run_log": []}
    col2key = dict(zip(ROLE_COLS, ROLE_KEYS))
    # legacy (pre two-axis) workbooks: map the old blended score onto both axes
    col2key.setdefault("Match /10", "fit_score")
    for r in read_sheet(wb["All Roles"]):
        row = {col2key.get(k, k): v for k, v in r.items() if k in col2key}
        if "attainability_score" not in row or row.get("attainability_score") in (None, ""):
            row["attainability_score"] = row.get("fit_score")
        data["roles"].append(row)
    if "Leads & Watchlist" in wb.sheetnames:
        data["leads"] = [{"type": r.get("Type"), "item": r.get("Item"), "details": r.get("Details"),
                          "action": r.get("Action"), "links": r.get("Link(s)")} for r in read_sheet(wb["Leads & Watchlist"])]
    if "Top Picks & Deadlines" in wb.sheetnames:
        data["top_picks"] = [{"list": r.get("List"), "rank": r.get("Rank"), "role_org": r.get("Role — Organisation"),
                              "match_score": r.get("Match /10"), "deadline": r.get("Deadline / urgency"), "link": r.get("Link")}
                             for r in read_sheet(wb["Top Picks & Deadlines"])]
    if "CV Keywords" in wb.sheetnames:
        data["cv_keywords"] = [{"category": r.get("Category"), "keywords": r.get("Keywords / guidance")} for r in read_sheet(wb["CV Keywords"])]
    if "Boards & Channels" in wb.sheetnames:
        data["boards"] = [{"cadence": r.get("Cadence"), "board": r.get("Board / channel"), "link": r.get("Link / note")} for r in read_sheet(wb["Boards & Channels"])]
    if "Coverage Report" in wb.sheetnames:
        data["coverage"] = [{"source": r.get("Source"), "attempted": r.get("Attempted"), "reachable": r.get("Reachable"),
                             "fetches": r.get("Fetches"), "roles_found": r.get("Found"), "roles_verified": r.get("Verified"),
                             "blocked_why": r.get("Blocked why"), "fix": r.get("Fix")} for r in read_sheet(wb["Coverage Report"])]
    if "Market Notes & Method" in wb.sheetnames:
        data["notes"] = [{"topic": r.get("Topic"), "notes": r.get("Notes")} for r in read_sheet(wb["Market Notes & Method"])]
    if "Run Log" in wb.sheetnames:
        data["run_log"] = [{"date": r.get("Date"), "roles_found": r.get("Roles found"), "new_added": r.get("New added"),
                            "expired": r.get("Expired"), "sources_nothing_new": r.get("Sources yielding nothing")} for r in read_sheet(wb["Run Log"])]
    return data


def main():
    args = sys.argv[1:]
    if len(args) < 2:
        sys.exit(__doc__)
    prior = load_prior(args[0])
    if args[1] == "--exclusions-only":
        for r in prior["roles"]:
            print(f"{r.get('organisation')} — {r.get('role_title')}")
        return
    new = json.load(open(args[1]))
    out = args[2]
    expired = []
    if "--expired" in args:
        expired = json.load(open(args[args.index("--expired") + 1]))
    today = new.get("meta", {}).get("run_date") or date.today().isoformat()

    pool = list(prior["roles"])
    added, skipped = [], []
    for r in new.get("roles", []):
        if any(same_role(r, p) for p in pool):
            skipped.append(f"{r.get('organisation')} — {r.get('role_title')}")
            continue
        r["run_date_added"] = today
        r.setdefault("source_report", f"Delta run {today}")
        added.append(r)
        pool.append(r)

    n_exp = 0
    for e in expired:
        for r in prior["roles"]:
            if same_role(e, r) and "expired" not in str(r.get("priority", "")).lower():
                r["priority"] = "Expired — kept for history"
                r["verification"] = f"{r.get('verification','')} | Expired {today}: {e.get('reason','closed')}"
                n_exp += 1

    merged = {
        "meta": {**new.get("meta", {}), "run_date": today},
        "roles": prior["roles"] + added,
        "leads": prior["leads"] + new.get("leads", []),
        "top_picks": new.get("top_picks") or prior["top_picks"],
        "cv_keywords": prior["cv_keywords"] + [k for k in new.get("cv_keywords", []) if k not in prior["cv_keywords"]],
        "boards": prior["boards"] + [b for b in new.get("boards", []) if b not in prior["boards"]],
        "notes": prior["notes"] + new.get("notes", []),
        "run_log": prior["run_log"] + [{"date": today, "roles_found": len(new.get("roles", [])), "new_added": len(added),
                                        "expired": n_exp, "sources_nothing_new": new.get("meta", {}).get("sources_nothing_new", "")}],
    }
    total = build(merged, out, run_date=today)
    print(json.dumps({"total_rows": total, "new_added": len(added), "duplicates_skipped": len(skipped),
                      "expired_marked": n_exp, "skipped": skipped}, indent=1))


if __name__ == "__main__":
    main()
