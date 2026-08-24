#!/usr/bin/env python3
"""Build the six-sheet master job-search workbook from a roles JSON file.

Usage:  python build_master_xlsx.py roles.json out.xlsx

roles.json shape: see references/output-schema.md. Only "roles" is required.
No formulas are written (pure data), so no recalculation step is needed.
"""
import json, re, sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROLE_COLS = ["No.", "Source report", "Scope", "Role title", "Organisation",
             "Location / arrangement", "Category", "Board / channel found on", "Posted",
             "Why it matches", "Requirements & key details", "Salary", "Deadline",
             "Direct application link", "Source link", "Fit /10", "Attainability /10",
             "Ladder", "Ghost status", "Priority", "Next action",
             "Verification status", "Flags", "Run date added"]
ROLE_KEYS = ["no", "source_report", "scope", "role_title", "organisation",
             "location_arrangement", "category", "board_found_on", "posted_date",
             "why_it_matches", "requirements", "salary", "deadline", "apply_link",
             "source_link", "fit_score", "attainability_score", "ladder",
             "ghost_status", "priority", "next_action", "verification", "flags",
             "run_date_added"]
ROLE_WIDTHS = [5, 20, 15, 28, 22, 30, 15, 22, 11, 42, 42, 12, 14, 42, 38, 8, 12, 7, 12, 16, 18, 30, 18, 12]

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name=ARIAL, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
APPLY_FILL = PatternFill("solid", fgColor="E2EFDA")
EXPIRED_FILL = PatternFill("solid", fgColor="EDEDED")


def demd(s):
    s = "" if s is None else str(s)
    s = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1 — \2", s)
    return s.replace("**", "").replace("`", "").strip()


def style(ws, widths, n_rows):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for c in ws[1]:
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, WRAP, THIN
    for r in range(2, n_rows + 2):
        for j in range(1, len(widths) + 1):
            c = ws.cell(row=r, column=j)
            c.font, c.alignment, c.border = CELL_FONT, WRAP, THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{max(n_rows, 1) + 1}"


def sort_key(r):
    try:
        f = float(str(r.get("fit_score", r.get("match_score", 0)) or 0).split()[0])
        at = float(str(r.get("attainability_score", f) or f).split()[0])
        score = (f + at) / 2
    except (ValueError, TypeError):
        score = 0
    dl = str(r.get("deadline", ""))
    urgent = 0 if re.search(r"\d{4}-\d{2}-\d{2}|\d{1,2} \w{3}", dl) else 1
    return (-score, urgent)


def add_sheet(wb, name, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append([demd(x) for x in r])
    style(ws, widths, len(rows))
    return ws


def build(data, out_path, run_date=None):
    run_date = run_date or data.get("meta", {}).get("run_date") or date.today().isoformat()
    roles = data.get("roles", [])
    for r in roles:
        r.setdefault("run_date_added", run_date)
        if isinstance(r.get("flags"), list):
            r["flags"] = ", ".join(r["flags"])
    roles = sorted(roles, key=sort_key)
    for i, r in enumerate(roles, 1):
        r["no"] = i

    wb = Workbook()
    ws = wb.active
    ws.title = "All Roles"
    ws.append(ROLE_COLS)
    for r in roles:
        ws.append([demd(r.get(k, "")) for k in ROLE_KEYS])
    style(ws, ROLE_WIDTHS, len(roles))
    for i, r in enumerate(roles, 2):
        pri = str(r.get("priority", "")).lower()
        fill = APPLY_FILL if "apply now" in pri else EXPIRED_FILL if "expired" in pri else None
        if fill:
            for j in range(1, len(ROLE_COLS) + 1):
                ws.cell(row=i, column=j).fill = fill

    add_sheet(wb, "Leads & Watchlist", ["Type", "Item", "Details", "Action", "Link(s)"],
              [[l.get("type"), l.get("item"), l.get("details"), l.get("action"), l.get("links")]
               for l in data.get("leads", [])], [16, 40, 70, 30, 55])
    add_sheet(wb, "Top Picks & Deadlines", ["List", "Rank", "Role — Organisation", "Match /10", "Deadline / urgency", "Link"],
              [[p.get("list"), p.get("rank"), p.get("role_org"), p.get("match_score"), p.get("deadline"), p.get("link")]
               for p in data.get("top_picks", [])], [24, 6, 62, 10, 20, 55])
    add_sheet(wb, "CV Keywords", ["Category", "Keywords / guidance"],
              [[k.get("category"), k.get("keywords")] for k in data.get("cv_keywords", [])], [38, 130])
    add_sheet(wb, "Boards & Channels", ["Cadence", "Board / channel", "Link / note"],
              [[b.get("cadence"), b.get("board"), b.get("link")] for b in data.get("boards", [])], [28, 60, 80])
    add_sheet(wb, "Coverage Report", ["Source", "Attempted", "Reachable", "Fetches", "Found", "Verified", "Blocked why", "Fix"],
              [[c.get("source"), c.get("attempted"), c.get("reachable"), c.get("fetches"),
                c.get("roles_found"), c.get("roles_verified"), c.get("blocked_why"), c.get("fix")]
               for c in data.get("coverage", [])], [30, 10, 14, 9, 8, 9, 45, 40])
    add_sheet(wb, "Market Notes & Method", ["Topic", "Notes"],
              [[n.get("topic"), n.get("notes")] for n in data.get("notes", [])], [30, 160])
    log = data.get("run_log", [{"date": run_date, "roles_found": len(roles), "new_added": len(roles),
                                "expired": 0, "sources_nothing_new": data.get("meta", {}).get("sources_nothing_new", "")}])
    add_sheet(wb, "Run Log", ["Date", "Roles found", "New added", "Expired", "Audit correction rate", "Sources yielding nothing"],
              [[e.get("date"), e.get("roles_found"), e.get("new_added"), e.get("expired"),
                e.get("audit_correction_rate", data.get("meta", {}).get("audit_correction_rate", "")),
                e.get("sources_nothing_new")] for e in log],
              [12, 12, 12, 10, 18, 80])
    wb.save(out_path)
    return len(roles)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("usage: build_master_xlsx.py roles.json out.xlsx")
    data = json.load(open(sys.argv[1]))
    n = build(data, sys.argv[2])
    print(f"wrote {sys.argv[2]} with {n} roles")
